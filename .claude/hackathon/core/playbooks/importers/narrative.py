"""Docusign-style narrative xlsx importer.

Each playbook sheet has columns:
  Section/Topic | Customer Issue/Request | Approvals/Notes/Responses | Escalation Owner(s) | Edits

Rows are grouped by Section/Topic header (column A). A header row contains text
in col A and 'Clause' or None in col B. Subsequent rows with col A blank are
sub-rules within that section. Each non-empty 'Customer Issue/Request' row
becomes one rule.
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional

import openpyxl

from ..store import PlaybookStore


HEADERS = ("section/topic", "customer issue/request",
           "approvals/notes/responses", "escalation owner(s)")


def _is_header(row: tuple) -> Optional[dict]:
    cells = [str(c).strip().lower() if c else "" for c in row]
    found = {c: i for i, c in enumerate(cells) if c in HEADERS}
    if len(found) >= 3:  # tolerate missing "Edits"
        return {
            "topic": found.get("section/topic"),
            "issue": found.get("customer issue/request"),
            "response": found.get("approvals/notes/responses"),
            "owner": found.get("escalation owner(s)"),
            "edits": next((i for i, c in enumerate(cells) if c.startswith("edits")), None),
        }
    return None


def import_xlsx(store: PlaybookStore, path: str, *,
                name: Optional[str] = None, owner_org: str = "Docusign") -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        pid = store.create_playbook(
            name=name or Path(path).stem, owner_org=owner_org, source_file=path
        )
        n_rules = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # locate header
            hmap = None
            header_idx = -1
            for idx, row in enumerate(rows[:6]):
                hmap = _is_header(row)
                if hmap:
                    header_idx = idx
                    break
            if not hmap:
                continue

            current_section = None
            for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                topic = row[hmap["topic"]] if hmap["topic"] is not None else None
                issue = row[hmap["issue"]] if hmap["issue"] is not None else None
                resp = row[hmap["response"]] if hmap["response"] is not None else None
                owner = row[hmap["owner"]] if hmap["owner"] is not None else None
                edits = row[hmap["edits"]] if hmap.get("edits") is not None else None

                if topic and not issue:  # section header
                    current_section = str(topic)
                    continue
                if not issue:
                    continue
                if str(issue).strip().lower() == "clause":
                    continue

                severity = "approval_required" if resp and "approval required" in str(resp).lower() else "warn"

                rid = store.create_rule(
                    playbook_id=pid,
                    title=f"{current_section or sheet_name} :: {str(issue)[:80]}",
                    description=str(issue),
                    applies_to="cluster",
                    severity=severity,
                    nl_assertion=str(issue),
                    preferred_language=str(edits) if edits else None,
                    escalation_owner=str(owner) if owner else None,
                    rationale=str(resp) if resp else None,
                    tags=[sheet_name, current_section or ""],
                    source_provenance={"file": path, "sheet": sheet_name, "row": ri},
                    status="draft",
                )
                if current_section:
                    store.add_binding(
                        rule_id=rid, entity_kind="cluster",
                        entity_id=f"label:{current_section}",
                        label_text=current_section, confidence=0.5,
                    )
                n_rules += 1

        store.conn.execute(
            "UPDATE playbooks SET description=? WHERE playbook_id=?",
            [f"Imported {n_rules} rules from {Path(path).name}", pid],
        )
        return pid
    finally:
        wb.close()
