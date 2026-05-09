"""Cloudera/Walmart tabular xlsx importer.

Header detection: contains Topic AND (Sub-Topic|Subtopic) AND
'Prompt Question'|'Subtopic Question' AND 'Answer Type'|'Answer Options'.
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional

import openpyxl

from ..store import PlaybookStore
from . import walmart_logic


def _header_map(row: tuple) -> Optional[dict]:
    cells = [str(c).strip() if c else "" for c in row]
    norm = [c.lower() for c in cells]

    def find(*names) -> Optional[int]:
        for i, c in enumerate(norm):
            if c in [n.lower() for n in names]:
                return i
        return None

    topic = find("topic")
    sub = find("sub-topic", "subtopic")
    q = find("prompt question", "subtopic question")
    a = find("answer type", "answer options")
    if None in (topic, sub, q, a):
        return None
    return {
        "topic": topic, "sub": sub, "question": q, "answer_type": a,
        "example": find("customer example languge", "customer example language"),
        "logic": find("deployment logic"),
    }


def _split_options(s: str) -> list[str]:
    if not s:
        return []
    return [t.strip() for t in s.replace(":", ";").split(";") if t.strip()]


def import_xlsx(store: PlaybookStore, path: str, *,
                name: Optional[str] = None,
                owner_org: str = "") -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        pid = store.create_playbook(
            name=name or Path(path).stem, owner_org=owner_org, source_file=path
        )

        n_rules = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # find header row
            hmap = None
            header_idx = -1
            for idx, row in enumerate(rows[:5]):
                hmap = _header_map(row)
                if hmap:
                    header_idx = idx
                    break
            if not hmap:
                continue

            for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                topic = row[hmap["topic"]] if hmap["topic"] is not None else None
                sub = row[hmap["sub"]] if hmap["sub"] is not None else None
                q = row[hmap["question"]] if hmap["question"] is not None else None
                if not (topic and sub and q):
                    continue
                ans_type_raw = row[hmap["answer_type"]] if hmap["answer_type"] is not None else ""
                options = _split_options(str(ans_type_raw or ""))
                example = row[hmap["example"]] if hmap.get("example") is not None else None
                logic = row[hmap["logic"]] if hmap.get("logic") is not None else None

                predicate = None
                if logic:
                    parsed = walmart_logic.parse(str(logic))
                    # Use the FIRST non-else branch as the rule's predicate; alternative branches
                    # (additional non-else clauses, and the else clause) are discarded for now.
                    primary = next((p for p in parsed if not p["is_else"]), None)
                    if primary:
                        predicate = primary["predicate"]

                rid = store.create_rule(
                    playbook_id=pid,
                    title=f"{topic} :: {sub}",
                    description=str(q),
                    applies_to="cluster",
                    severity="warn",
                    predicate=predicate,
                    reference_text=str(example) if example else None,
                    answer_type="standard_nonstandard" if "standard" in str(ans_type_raw).lower() else "yes_no",
                    answer_options=options,
                    tags=[str(topic), str(sub)],
                    source_provenance={"file": path, "sheet": sheet_name, "row": ri},
                    status="draft",
                )
                store.add_binding(
                    rule_id=rid, entity_kind="cluster",
                    entity_id=f"label:{topic}",   # soft binding placeholder; resolver fills cluster_id
                    label_text=str(topic), confidence=0.5,
                )
                n_rules += 1

        store.conn.execute(
            "UPDATE playbooks SET description=? WHERE playbook_id=?",
            [f"Imported {n_rules} rules from {Path(path).name}", pid],
        )
        return pid
    finally:
        wb.close()
